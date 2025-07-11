#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Teste Final - Verificação completa do openpyxl e sistema
"""
import sys
import traceback
from datetime import datetime

def teste_openpyxl():
    """Teste completo da funcionalidade openpyxl"""
    print("🧪 TESTE: openpyxl")
    print("=" * 50)
    
    try:
        # Teste 1: Import básico
        import openpyxl
        print(f"✅ Import openpyxl: OK (versão {openpyxl.__version__})")
        
        # Teste 2: Criar workbook
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Teste"
        print("✅ Criação de Workbook: OK")
        
        # Teste 3: Adicionar dados
        ws['A1'] = "Teste"
        ws['B1'] = 123.45
        ws['C1'] = datetime.now()
        print("✅ Adição de dados: OK")
        
        # Teste 4: Salvar arquivo
        wb.save("teste_openpyxl.xlsx")
        print("✅ Salvamento de arquivo: OK")
        
        # Teste 5: Pandas ExcelWriter
        import pandas as pd
        df = pd.DataFrame({
            'Nome': ['João', 'Maria', 'Pedro'],
            'Idade': [25, 30, 35],
            'Salário': [3000.50, 4500.75, 5200.00]
        })
        
        with pd.ExcelWriter('teste_pandas_excel.xlsx', engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Funcionarios', index=False)
        print("✅ Pandas ExcelWriter: OK")
        
        return True
        
    except Exception as e:
        print(f"❌ ERRO no teste openpyxl: {e}")
        traceback.print_exc()
        return False

def teste_sistema_dependencias():
    """Teste das principais dependências do sistema"""
    print("\n🧪 TESTE: Dependências do Sistema")
    print("=" * 50)
    
    dependencias = [
        ('numpy', 'np'),
        ('pandas', 'pd'),
        ('scipy', None),
        ('sklearn', None),  # Changed from 'scikit-learn' to 'sklearn'
        ('matplotlib.pyplot', 'plt'),
        ('arch', None),
        ('statsmodels.api', 'sm'),
        ('tensorflow', 'tf'),
        ('keras', None),
        ('MetaTrader5', 'mt5'),
        ('plotly', None)
    ]
    
    sucessos = 0
    falhas = 0
    
    for nome, alias in dependencias:
        try:
            if alias:
                exec(f"import {nome} as {alias}")
            else:
                exec(f"import {nome}")
            print(f"✅ {nome}: OK")
            sucessos += 1
        except ImportError as e:
            print(f"❌ {nome}: FALHA - {e}")
            falhas += 1
        except Exception as e:
            print(f"⚠️ {nome}: ERRO - {e}")
            falhas += 1
    
    print(f"\n📊 RESUMO: {sucessos} sucessos, {falhas} falhas")
    return falhas == 0

def main():
    """Função principal do teste"""
    print("🎯 VERIFICAÇÃO FINAL - OPENPYXL E DEPENDÊNCIAS")
    print("=" * 60)
    print(f"Data/Hora: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"Python: {sys.version}")
    print()
    
    # Teste openpyxl
    openpyxl_ok = teste_openpyxl()
    
    # Teste dependências
    dependencias_ok = teste_sistema_dependencias()
    
    # Resultado final
    print("\n🏆 RESULTADO FINAL")
    print("=" * 50)
    
    if openpyxl_ok and dependencias_ok:
        print("✅ TODOS OS TESTES PASSARAM!")
        print("✅ Sistema pronto para execução!")
        status = "SUCESSO"
    else:
        print("❌ ALGUNS TESTES FALHARAM!")
        if not openpyxl_ok:
            print("❌ openpyxl com problemas")
        if not dependencias_ok:
            print("❌ Dependências com problemas")
        status = "FALHA"
    
    print(f"\n📋 Status final: {status}")
    return status == "SUCESSO"

if __name__ == "__main__":
    main()
